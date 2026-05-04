# -*- coding: utf-8 -*-
"""
Flask 主应用模块
提供所有 API 路由、爬虫任务管理、Excel 导出等功能
"""

import os
import sys
import uuid
import logging
import threading
from datetime import datetime
from io import BytesIO

from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 将项目根目录加入 Python 路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import FLASK_CONFIG, DEFAULT_PER_PAGE, MAX_PER_PAGE, EXPORT_CONFIG
from models import init_database, query_jobs, get_job_by_id, get_all_companies, get_all_industries, get_stats, export_jobs, batch_insert_jobs

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs', 'app.log'), encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

# 创建 Flask 应用
app = Flask(__name__)

# ==================== 全局爬虫状态管理 ====================
# 使用线程锁保护状态字典的并发访问
crawl_status_lock = threading.Lock()

# 爬虫状态字典
crawl_status = {
    'status': 'idle',       # idle / running / done / error
    'progress': '',          # 当前进度描述
    'total_new': 0,          # 新增数据条数
    'error_message': '',     # 错误信息
    'task_id': '',           # 任务 ID
    'source': '',            # 当前抓取的数据源
}


def update_crawl_status(status=None, progress=None, total_new=None, error_message=None, task_id=None, source=None):
    """
    更新爬虫状态（线程安全）
    
    Args:
        status: 状态值
        progress: 进度描述
        total_new: 新增数据条数
        error_message: 错误信息
        task_id: 任务 ID
        source: 数据源
    """
    with crawl_status_lock:
        if status is not None:
            crawl_status['status'] = status
        if progress is not None:
            crawl_status['progress'] = progress
        if total_new is not None:
            crawl_status['total_new'] = total_new
        if error_message is not None:
            crawl_status['error_message'] = error_message
        if task_id is not None:
            crawl_status['task_id'] = task_id
        if source is not None:
            crawl_status['source'] = source


def run_crawler_task(source):
    """
    在后台线程中执行爬虫任务
    爬虫每抓到一批数据就实时推送到数据库

    Args:
        source: 数据源类型 (shixiseng / ncss / websites / all)
    """
    try:
        update_crawl_status(
            status='running',
            progress=f'正在启动 {source} 爬虫...',
            total_new=0,
            error_message='',
            source=source
        )

        total_new = [0]

        def progress_callback(message, count):
            update_crawl_status(progress=message, total_new=total_new[0])

        def on_jobs_found(jobs_list):
            """实时回调：爬虫每抓到一批数据就调用此函数"""
            inserted, updated = batch_insert_jobs(jobs_list)
            total_new[0] += inserted
            update_crawl_status(total_new=total_new[0])
            logger.info(f"实时插入: 新增 {inserted} 条, 更新 {updated} 条, 累计 {total_new[0]} 条")

        if source in ('shixiseng', 'all'):
            update_crawl_status(progress='正在抓取实习僧数据...')
            try:
                from crawler.shixiseng_crawler import ShixisengCrawler
                crawler = ShixisengCrawler()
                crawler.set_progress_callback(progress_callback)
                crawler.set_jobs_callback(on_jobs_found)
                jobs = crawler.run()
                # 获取详情页的薪资和描述
                if jobs:
                    update_crawl_status(progress='实习僧 - 正在获取岗位详情...')
                    crawler.start_browser()
                    crawler.enrich_jobs(jobs, max_count=100)
                    crawler.close_browser()
                    # 更新数据库中的详情信息
                    from models import batch_insert_jobs as batch_update
                    batch_update(jobs)
                logger.info(f"实习僧爬虫结束，共处理 {len(jobs)} 条数据")
            except Exception as e:
                logger.error(f"实习僧爬虫异常: {e}")
                update_crawl_status(progress=f'实习僧抓取出错: {str(e)}')

        if source in ('ncss', 'all'):
            update_crawl_status(progress='正在抓取国家平台数据...')
            try:
                from crawler.ncss_crawler import NCSSCrawler
                crawler = NCSSCrawler()
                crawler.set_progress_callback(progress_callback)
                crawler.set_jobs_callback(on_jobs_found)
                jobs = crawler.run()
                logger.info(f"国家平台爬虫结束，共处理 {len(jobs)} 条数据")
            except Exception as e:
                logger.error(f"国家平台爬虫异常: {e}")
                update_crawl_status(progress=f'国家平台抓取出错: {str(e)}')

        if source in ('websites', 'all'):
            update_crawl_status(progress='正在抓取大厂官网数据...')
            try:
                from crawler.website_crawler import WebsiteCrawler
                crawler = WebsiteCrawler()
                crawler.set_progress_callback(progress_callback)
                crawler.set_jobs_callback(on_jobs_found)
                jobs = crawler.run()
                logger.info(f"大厂官网爬虫结束，共处理 {len(jobs)} 条数据")
            except Exception as e:
                logger.error(f"大厂官网爬虫异常: {e}")
                update_crawl_status(progress=f'大厂官网抓取出错: {str(e)}')

        update_crawl_status(
            status='done',
            progress=f'抓取完成！共新增 {total_new[0]} 条数据',
            total_new=total_new[0]
        )
        logger.info(f"爬虫任务完成，共新增 {total_new[0]} 条数据")

    except Exception as e:
        logger.error(f"爬虫任务异常: {e}")
        update_crawl_status(
            status='error',
            progress='抓取出错',
            error_message=str(e)
        )


# ==================== 页面路由 ====================

@app.route('/')
def index():
    """渲染首页"""
    return render_template('index.html')


# ==================== API 路由 ====================

@app.route('/api/jobs', methods=['GET'])
def api_get_jobs():
    """
    获取岗位列表 API
    
    查询参数：
    - keyword: 岗位关键词
    - company: 公司名称
    - location: 工作地点
    - job_type: 岗位类型 (intern/graduate/all)
    - salary_min: 最低薪资
    - salary_max: 最高薪资
    - date_from: 开始日期
    - date_to: 结束日期
    - source: 数据来源（逗号分隔）
    - page: 页码（默认1）
    - per_page: 每页条数（默认20）
    """
    try:
        # 获取查询参数
        keyword = request.args.get('keyword', '').strip()
        company = request.args.get('company', '').strip()
        location = request.args.get('location', '').strip()
        job_type = request.args.get('job_type', 'all').strip()
        salary_min = request.args.get('salary_min', type=int)
        salary_max = request.args.get('salary_max', type=int)
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()
        source = request.args.get('source', '').strip()
        industry = request.args.get('industry', '').strip()
        company_nature = request.args.get('company_nature', '').strip()
        education = request.args.get('education', '').strip()
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', DEFAULT_PER_PAGE, type=int)

        # 限制每页条数
        per_page = min(per_page, MAX_PER_PAGE)
        per_page = max(per_page, 1)
        page = max(page, 1)

        # 查询数据
        result = query_jobs(
            keyword=keyword,
            company=company,
            location=location,
            job_type=job_type,
            salary_min=salary_min,
            salary_max=salary_max,
            date_from=date_from,
            date_to=date_to,
            source=source,
            industry=industry,
            company_nature=company_nature,
            education=education,
            page=page,
            per_page=per_page
        )
        
        return jsonify(result)
    
    except Exception as e:
        logger.error(f"获取岗位列表失败: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/jobs/<int:job_id>', methods=['GET'])
def api_get_job(job_id):
    """
    获取单个岗位详情 API
    
    Args:
        job_id: 岗位 ID
    """
    try:
        job = get_job_by_id(job_id)
        if job:
            return jsonify(job)
        else:
            return jsonify({'error': '岗位不存在'}), 404
    except Exception as e:
        logger.error(f"获取岗位详情失败: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/crawl', methods=['POST'])
def api_start_crawl():
    """
    启动爬虫任务 API
    
    请求体 JSON:
    - source: 数据源 (shixiseng / ncss / websites / all)
    """
    try:
        data = request.get_json() or {}
        source = data.get('source', 'all')
        
        # 验证数据源参数
        valid_sources = ('shixiseng', 'ncss', 'websites', 'all')
        if source not in valid_sources:
            return jsonify({'error': f'无效的数据源: {source}，有效值: {valid_sources}'}), 400
        
        # 检查是否已有任务在运行
        with crawl_status_lock:
            if crawl_status['status'] == 'running':
                return jsonify({'error': '已有爬虫任务正在运行，请等待完成'}), 409
        
        # 生成任务 ID
        task_id = str(uuid.uuid4())
        
        # 更新状态为运行中
        update_crawl_status(
            status='running',
            progress='正在启动爬虫任务...',
            total_new=0,
            error_message='',
            task_id=task_id,
            source=source
        )
        
        # 在后台线程中执行爬虫任务
        thread = threading.Thread(target=run_crawler_task, args=(source,), daemon=True)
        thread.start()
        
        logger.info(f"爬虫任务已启动: task_id={task_id}, source={source}")
        
        return jsonify({
            'task_id': task_id,
            'status': 'started'
        })
    
    except Exception as e:
        logger.error(f"启动爬虫任务失败: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/crawl/status', methods=['GET'])
def api_crawl_status():
    """
    获取爬虫任务状态 API
    
    返回 JSON:
    - status: 状态 (idle/running/done/error)
    - progress: 进度描述
    - total_new: 新增数据条数
    - error_message: 错误信息
    """
    with crawl_status_lock:
        return jsonify({
            'status': crawl_status['status'],
            'progress': crawl_status['progress'],
            'total_new': crawl_status['total_new'],
            'error_message': crawl_status['error_message'],
        })


@app.route('/api/export', methods=['GET'])
def api_export_excel():
    """
    导出 Excel API
    
    接受与 /api/jobs 完全相同的筛选参数
    生成 Excel 文件并返回下载
    """
    try:
        # 获取筛选参数
        keyword = request.args.get('keyword', '').strip()
        company = request.args.get('company', '').strip()
        location = request.args.get('location', '').strip()
        job_type = request.args.get('job_type', 'all').strip()
        salary_min = request.args.get('salary_min', type=int)
        salary_max = request.args.get('salary_max', type=int)
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()
        source = request.args.get('source', '').strip()
        industry = request.args.get('industry', '').strip()
        company_nature = request.args.get('company_nature', '').strip()
        education = request.args.get('education', '').strip()

        # 查询数据（不分页）
        jobs = export_jobs(
            keyword=keyword,
            company=company,
            location=location,
            job_type=job_type,
            salary_min=salary_min,
            salary_max=salary_max,
            date_from=date_from,
            date_to=date_to,
            source=source,
            industry=industry,
            company_nature=company_nature,
            education=education
        )
        
        if not jobs:
            return jsonify({'error': '没有符合条件的数据可导出'}), 404
        
        # 限制导出行数
        if len(jobs) > EXPORT_CONFIG['max_rows']:
            jobs = jobs[:EXPORT_CONFIG['max_rows']]
        
        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = EXPORT_CONFIG['sheet_name']
        
        # 定义表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 定义表头
        headers = ['序号', '岗位名称', '公司名称', '工作地点', '薪资', '岗位类型', '学历要求', '行业', '公司性质', '公司规模', '福利', '发布时间', '岗位描述', '详情链接', '数据来源']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 填充数据
        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(vertical='center', wrap_text=True)
        
        for idx, job in enumerate(jobs, 1):
            row = idx + 1
            
            # 岗位类型中文转换
            job_type_cn = '实习' if job.get('job_type') == 'intern' else '校招'
            
            row_data = [
                idx,
                job.get('job_title', ''),
                job.get('company_name', ''),
                job.get('location', ''),
                job.get('salary', ''),
                job_type_cn,
                job.get('education', ''),
                job.get('industry', ''),
                job.get('company_nature', ''),
                job.get('company_size', ''),
                job.get('welfare', ''),
                job.get('publish_date', ''),
                job.get('job_desc', '')[:500],  # 限制描述长度
                job.get('job_url', ''),
                job.get('source', ''),
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
        
        # 设置列宽
        column_widths = [6, 30, 20, 15, 12, 10, 10, 18, 10, 12, 25, 12, 50, 40, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = width
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        now_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        filter_summary = ''
        if keyword:
            filter_summary += f'_{keyword}'
        if company:
            filter_summary += f'_{company}'
        if source:
            filter_summary += f'_{source}'
        if not filter_summary:
            filter_summary = '_全部'
        
        filename = f'岗位数据{filter_summary}_{now_str}.xlsx'
        
        logger.info(f"导出 Excel: {filename}, 共 {len(jobs)} 条数据")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        logger.error(f"导出 Excel 失败: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/companies', methods=['GET'])
def api_get_companies():
    """
    获取所有公司名称列表 API
    用于前端搜索下拉选择
    """
    try:
        companies = get_all_companies()
        return jsonify(companies)
    except Exception as e:
        logger.error(f"获取公司列表失败: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/industries', methods=['GET'])
def api_get_industries():
    """
    获取所有行业列表 API
    用于前端行业筛选下拉
    """
    try:
        industries = get_all_industries()
        return jsonify(industries)
    except Exception as e:
        logger.error(f"获取行业列表失败: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/stats', methods=['GET'])
def api_get_stats():
    """
    获取统计数据 API
    
    返回 JSON:
    - total_jobs: 总岗位数
    - sources: 各来源数量
    - last_updated: 最后更新时间
    """
    try:
        stats = get_stats()
        return jsonify(stats)
    except Exception as e:
        logger.error(f"获取统计数据失败: {e}")
        return jsonify({'error': str(e)}), 500


# ==================== 应用启动 ====================

def create_app():
    """
    创建并配置 Flask 应用
    
    Returns:
        Flask: 配置好的 Flask 应用实例
    """
    # 初始化数据库
    init_database()
    logger.info("数据库初始化完成")
    
    return app


if __name__ == '__main__':
    # 创建应用
    app = create_app()
    
    # 启动应用
    logger.info(f"应用启动: http://{FLASK_CONFIG['host']}:{FLASK_CONFIG['port']}")
    app.run(
        host=FLASK_CONFIG['host'],
        port=FLASK_CONFIG['port'],
        debug=FLASK_CONFIG['debug']
    )
